"""
Llenador automático del Packing List C&A (entregas nacionales).

Toma los datos ya extraídos de una OC (ver cya_parser.py) más la capacidad
física de caja (packs por caja / piezas por caja -- esto SIEMPRE lo defines
tú a mano porque depende del volumen de la prenda) y llena PL_ACTUALIZADO.xlsx
respetando exactamente las fórmulas y el formato del template original.
"""
from openpyxl import load_workbook
import copy


def _split_por_capacidad(total, capacidad):
    """Divide un total de packs/piezas en renglones de caja llena + caja resto.

    Devuelve lista de tuplas (cantidad_en_esa_caja, cajas_de_ese_tipo)
    -> en el Excel: (Packs per carton, Number of cartons) o
                    (Pieces per carton, Number of cartons)
    Si no se da capacidad, regresa un solo renglón vacío para llenado manual.
    """
    if not capacidad or capacidad <= 0:
        return [(None, None)]
    llenas = total // capacidad
    resto = total % capacidad
    filas = []
    if llenas > 0:
        filas.append((capacidad, llenas))       # (piezas/packs por caja, num de cajas)
    if resto > 0:
        filas.append((resto, 1))                # caja resto: siempre menor distribución
    if not filas:
        filas.append((0, 0))
    return filas


def _limpiar_fila(ws, fila, col_ini=3, col_fin=15):
    """Borra cualquier valor residual del template (ej. tallas '11-12' de un
    ejemplo previo) en el rango de columnas de tallas, antes de escribir los
    valores reales de esta OC. Evita que sobrevivan celdas viejas cuando la
    OC trae menos tallas que las que el template tenía de ejemplo.

    OJO: ws.cell(row=r, column=c, value=None) NO limpia nada -- openpyxl
    ignora value=None y deja la celda intacta. Hay que asignar .value
    directamente sobre el objeto celda.
    """
    for col in range(col_ini, col_fin + 1):
        ws.cell(row=fila, column=col).value = None


def fill_packing_list(template_path, output_path, oc_data,
                       capacidad_ratio=None, capacidad_solid=None,
                       cantidades_reales_solid=None,
                       shipper_code='70135', shipper_name_override=None,
                       gross_weight=None, net_weight=None, cbm=None,
                       invoice_number=None):
    """
    template_path: ruta a PL_ACTUALIZADO.xlsx (plantilla en blanco)
    output_path: ruta del archivo a generar
    oc_data: dict devuelto por cya_parser.parse_oc_pdf
    capacidad_ratio: dict {'A': packs_por_caja, 'B': packs_por_caja} (opcional)
    capacidad_solid: int (piezas por caja, un solo valor para todas las tallas)
                      o dict {talla: piezas_por_caja} si varía por talla
    cantidades_reales_solid: dict opcional {talla: piezas_reales_a_entregar}
      para el bloque Solid Pack (remanente). Úsalo cuando por faltantes de
      fabricación la cantidad real disponible es menor a lo planeado en la OC.
      Si no se pasa, se usa la cantidad de la OC tal cual (sin faltante).
      OJO: "Pieces per PO" (columna F) SIEMPRE queda con el valor original
      de la OC para ese SKU dentro de este pack -- nunca se toca, es la
      referencia de lo planeado, aunque E (Pieces to Delivery) cambie.
    shipper_code, gross_weight, net_weight, cbm, invoice_number: campos que
      NO vienen en la OC -- captúralos tú; si no se pasan, quedan en blanco
      para que los llenes a mano en el Excel resultante.
    """
    wb = load_workbook(template_path)
    ws = wb['Sheet1']
    header = oc_data['header']
    sku_table = oc_data['sku_table']
    sku_order = oc_data['sku_order']
    packs = oc_data['packs']

    ratio_packs = [p for p in packs if p['tipo'] == 'PACK']
    solid_packs = [p for p in packs if p['tipo'] == 'SKU']

    # --- Encabezado ---
    ws['B2'] = shipper_name_override or header.get('proveedor')
    if shipper_code:
        ws['B3'] = shipper_code
    ws['B4'] = header.get('modelo_id')
    ws['B5'] = header.get('division')
    ws['B6'] = header.get('numero_orden')
    if gross_weight is not None:
        ws['B9'] = gross_weight
    if net_weight is not None:
        ws['B10'] = net_weight
    if cbm is not None:
        ws['B11'] = cbm
    if invoice_number:
        ws['B12'] = invoice_number
    # B7 (TOTAL CARTONS) y B8 (TOTAL PIECES) son fórmulas -> no se tocan

    warnings = []

    # El color SIEMPRE se escribe en B18 (Pack A), sin importar si esta OC trae
    # Ratio Packs o no -- B23 y B36 lo heredan por fórmula (=+B18, =+B23), así
    # que si no se escribe aquí, se queda pegado el color de ejemplo del template.
    _limpiar_fila(ws, 18, col_ini=2, col_fin=2)
    ws.cell(row=18, column=2, value=header.get('color_generico'))  # B18

    # --- Bloques "Ratio Pack" (A: filas 18-19, B: filas 23-24) ---
    # Las tallas (fila 18 y 23) SIEMPRE se escriben con el catálogo de tallas
    # de la OC (sku_order) -- igual que en Pack C -- exista o no un Ratio Pack
    # real en esta OC. Si no hay Ratio Pack para esa letra, la fila de
    # cantidades (19/24) simplemente se queda en blanco.
    ratio_rows = {'A': (18, 19), 'B': (23, 24)}
    ratio_por_letra = {p['letra']: p for p in ratio_packs[:2]}
    for letra, (row_label, row_qty) in ratio_rows.items():
        _limpiar_fila(ws, row_label)
        _limpiar_fila(ws, row_qty)
        for col, talla in enumerate(sku_order, start=3):
            ws.cell(row=row_label, column=col, value=talla)

    for pack in ratio_packs[:2]:  # el template solo soporta A y B
        letra = pack['letra']
        if letra not in ratio_rows:
            warnings.append(f"Pack '{letra}' no tiene bloque Ratio Pack disponible en el template (solo A y B).")
            continue
        _, row_qty = ratio_rows[letra]
        for col, talla in enumerate(sku_order, start=3):
            if talla in pack['tallas']:
                cantidad_total = pack['tallas'][talla]
                total_packs = pack['total_packs']
                ratio = cantidad_total / total_packs if total_packs else 0
                if ratio != int(ratio):
                    warnings.append(
                        f"Pack {letra}, talla {talla}: la razón {cantidad_total}/{total_packs} "
                        f"no da un entero exacto ({ratio}). Verifica el prepack."
                    )
                ws.cell(row=row_qty, column=col, value=round(ratio))

    # --- Tabla "PACK No." (filas 27-30): renglones CONTIGUOS, sin huecos.
    # Cada pack usa 1 renglón (si cabe en una sola caja) o 2 (caja llena + caja
    # resto) -- pero el siguiente pack empieza inmediatamente después, nunca
    # se deja un renglón vacío reservado.
    fila_pack_no = 27
    max_fila_pack_no = 30
    for pack in ratio_packs[:2]:
        letra = pack['letra']
        cap = (capacidad_ratio or {}).get(letra)
        splits = _split_por_capacidad(pack['total_packs'], cap)
        for packs_por_caja, num_cajas in splits:
            if fila_pack_no > max_fila_pack_no:
                warnings.append(
                    "Se agotaron los renglones disponibles en la tabla PACK No. (27-30); "
                    "faltó capturar alguna caja resto manualmente."
                )
                break
            ws.cell(row=fila_pack_no, column=1, value=pack['pack_id'])        # A: PACK No.
            ws.cell(row=fila_pack_no, column=2, value=letra)                  # B: PACK
            ws.cell(row=fila_pack_no, column=3, value=num_cajas)              # C: Number of cartons
            ws.cell(row=fila_pack_no, column=4, value=packs_por_caja)         # D: Packs per carton
            ws.cell(row=fila_pack_no, column=5, value=pack['unidades_por_pack'])  # E: Pieces Per Pack
            ws.cell(row=fila_pack_no, column=6, value=pack['total_packs'])    # F: Packs per PO
            # G ya trae la fórmula =C*D*E
            fila_pack_no += 1

    # Limpia cualquier renglón sobrante que el template haya dejado precargado
    # (ej. la letra 'B' de un pack que ya no llegó a usar ese renglón).
    for fila in range(fila_pack_no, max_fila_pack_no + 1):
        for col in range(1, 7):
            ws.cell(row=fila, column=col).value = None

    # --- Bloque "Solid Pack" (filas 36-37): distribución agregada por talla ---
    # La letra de este bloque es la que la OC realmente le puso al pack tipo SKU
    # (puede ser "B" si solo hay un Ratio Pack antes, "C" si hay dos, etc.) --
    # no siempre es "C". Si la OC no trae ningún desglose de Pack (Escenario 1),
    # se usa "C" por default ya que no hay letra real de la OC que tomar.
    solid_letra = solid_packs[0]['letra'] if solid_packs else 'C'

    # "planeado" = lo que la OC asignó a este pack tipo SKU para esa talla (fijo,
    #   es la referencia -- va en la columna F "Pieces per PO").
    # "a_entregar" = lo que realmente se va a mandar (editable vía
    #   cantidades_reales_solid, por defecto igual a lo planeado si no hay faltante).
    # Si no hay packs tipo SKU en la OC (Escenario 1), se usa la tabla SKU completa como "solid".
    if solid_packs:
        planeado_talla = {}
        for p in solid_packs:
            for talla, cant in p['tallas'].items():
                planeado_talla[talla] = planeado_talla.get(talla, 0) + cant
    else:
        planeado_talla = {t: sku_table[t]['piezas'] for t in sku_order}

    cantidades_reales_solid = cantidades_reales_solid or {}
    a_entregar_talla = {
        talla: cantidades_reales_solid.get(talla, planeado)
        for talla, planeado in planeado_talla.items()
    }

    _limpiar_fila(ws, 36, col_ini=3, col_fin=15)  # OJO: no tocar columna B (36), trae la fórmula del color
    ws.cell(row=36, column=1).value = None
    _limpiar_fila(ws, 37)
    ws.cell(row=36, column=1, value=solid_letra)  # A36: letra del pack (según la OC)
    for col, talla in enumerate(sku_order, start=3):
        ws.cell(row=36, column=col, value=talla)
        if talla in a_entregar_talla:
            ws.cell(row=37, column=col, value=a_entregar_talla[talla])  # cantidad REAL a entregar

    # --- Tabla SKU (Solid Pack), filas 40 en adelante ---
    fila = 40
    max_fila_sku = 55  # filas 56-58 son "MIXED", no aplican a entregas nacionales
    for talla in sku_order:
        if talla not in a_entregar_talla:
            continue
        cantidad_a_entregar = a_entregar_talla[talla]   # E: Pieces to Delivery (real)
        piezas_po = planeado_talla[talla]                # F: Pieces per PO (lo planeado en la OC para ESTE pack, fijo)
        sku_no = sku_table[talla]['sku']

        cap = capacidad_solid
        if isinstance(capacidad_solid, dict):
            cap = capacidad_solid.get(talla)
        splits = _split_por_capacidad(cantidad_a_entregar, cap)

        for pieces_por_caja, num_cajas in splits:
            if fila > max_fila_sku:
                warnings.append(
                    "Se agotaron los renglones disponibles en la tabla SKU (40-55); "
                    "faltó capturar algunas cajas resto manualmente."
                )
                break
            ws.cell(row=fila, column=1, value=sku_no)              # A: SKU No.
            ws.cell(row=fila, column=2, value=solid_letra)         # B: PACK (letra real de la OC)
            ws.cell(row=fila, column=3, value=num_cajas)           # C: Number of cartons
            ws.cell(row=fila, column=4, value=pieces_por_caja)     # D: Pieces per carton
            ws.cell(row=fila, column=5, value=cantidad_a_entregar) # E: Pieces to Delivery
            ws.cell(row=fila, column=6, value=piezas_po)           # F: Pieces per PO
            # G ya trae la fórmula =C*D
            fila += 1

    # Limpia el 'C' precargado del template en renglones 40-44 que no se usaron
    # (por ejemplo si el remanente cupo en menos renglones de los reservados).
    for fila_sobrante in range(fila, 45):
        for col in range(1, 7):
            ws.cell(row=fila_sobrante, column=col).value = None

    wb.save(output_path)
    return warnings
